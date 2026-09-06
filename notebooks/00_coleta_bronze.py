# Databricks notebook source
# =============================================================================
# MVP Engenharia de Dados — Ocorrências registradas em Sorocaba
# Notebook 00 — Coleta manual e camada Bronze
#
# Pré-requisito: os quatro XLSX originais da SSP-SP (2022–2025) devem ser
# enviados, sem alteração, para o Unity Catalog Volume configurado abaixo.
# Este notebook não acessa a internet, não agenda execuções e não implementa
# carga incremental. Toda execução reconstrói o recorte Bronze e seu manifesto.
# =============================================================================

# COMMAND ----------
# MAGIC %pip install openpyxl==3.1.5

# COMMAND ----------
# MAGIC %md
# MAGIC ## 1. Configuração e verificação dos arquivos originais
# MAGIC
# MAGIC Faça o upload manual de `SPDadosCriminais_2022.xlsx` a
# MAGIC `SPDadosCriminais_2025.xlsx` para o diretório exibido por esta célula.
# MAGIC Os arquivos permanecem imutáveis no Volume; a tabela Bronze contém apenas
# MAGIC o recorte de Sorocaba para tornar a execução compatível com o Free Edition.

# COMMAND ----------

import datetime as dt
import hashlib
import json
import os
import re
import unicodedata
from decimal import Decimal, InvalidOperation

import openpyxl
from pyspark.sql import functions as F
from pyspark.sql.types import (
    IntegerType,
    StringType,
    StructField,
    StructType,
    TimestampType,
)

CATALOG = "workspace"
SCHEMA = "sorocaba_seguranca"
VOLUME = "dados"
DIRETORIO_XLSX = f"/Volumes/{CATALOG}/{SCHEMA}/{VOLUME}/xlsx"
COD_IBGE_SOROCABA = "3552205"
LOTE_SOROCABA = 25_000

URL_TEMPLATE = (
    "https://www.ssp.sp.gov.br/assets/estatistica/transparencia/"
    "spDados/SPDadosCriminais_{ano}.xlsx"
)

CONFIG_ANOS = {
    2022: ("SPDadosCriminais_2022.xlsx", ["JAN-JUN_2022", "JUL-DEZ_2022"]),
    2023: ("SPDadosCriminais_2023.xlsx", ["JAN-JUN_2023", "JUL-DEZ_2023"]),
    2024: ("SPDadosCriminais_2024.xlsx", ["JAN-JUN_2024", "JUL-DEZ_2024"]),
    2025: ("SPDadosCriminais_2025.xlsx", ["JAN-JUN_2025", "JUL-DEZ_2025"]),
}

spark.sql(f"USE CATALOG {CATALOG}")
spark.sql(f"CREATE SCHEMA IF NOT EXISTS {CATALOG}.{SCHEMA}")
spark.sql(f"CREATE VOLUME IF NOT EXISTS {CATALOG}.{SCHEMA}.{VOLUME}")
spark.sql(f"USE {CATALOG}.{SCHEMA}")

os.makedirs(DIRETORIO_XLSX, exist_ok=True)

arquivos_esperados = [
    os.path.join(DIRETORIO_XLSX, arquivo)
    for arquivo, _ in CONFIG_ANOS.values()
]
arquivos_ausentes = [caminho for caminho in arquivos_esperados if not os.path.isfile(caminho)]

print(f"Catálogo : {CATALOG}")
print(f"Schema   : {SCHEMA}")
print(f"Volume   : {DIRETORIO_XLSX}")
print("Arquivos esperados:")
for caminho in arquivos_esperados:
    situacao = "OK" if os.path.isfile(caminho) else "AUSENTE"
    print(f"  [{situacao}] {os.path.basename(caminho)}")

if arquivos_ausentes:
    nomes = ", ".join(os.path.basename(caminho) for caminho in arquivos_ausentes)
    raise FileNotFoundError(
        "Faça o upload manual dos quatro XLSX originais para o Volume antes "
        f"de continuar. Arquivos ausentes: {nomes}"
    )

# COMMAND ----------
# MAGIC %md
# MAGIC ## 2. Funções de preservação, filtro e auditoria
# MAGIC
# MAGIC A leitura usa `openpyxl` no modo `read_only`. Todas as linhas estaduais
# MAGIC são percorridas, mas somente o código IBGE `3552205` chega ao Spark. Os valores
# MAGIC originais são serializados como texto; apenas os nomes das colunas são
# MAGIC normalizados para identificadores compatíveis com Delta Lake.

# COMMAND ----------


def renderizar_valor(valor):
    """Serializa uma célula sem aplicar regra de negócio ou limpeza."""
    if valor is None:
        return None
    if isinstance(valor, dt.datetime):
        return valor.isoformat(sep=" ", timespec="seconds")
    if isinstance(valor, dt.date):
        return valor.isoformat()
    if isinstance(valor, dt.time):
        return valor.isoformat(timespec="seconds")
    return str(valor)


def normalizar_identificador(nome, indice):
    """Converte o cabeçalho em identificador ASCII estável para Delta Lake."""
    if nome is None or not str(nome).strip():
        return f"COLUNA_SEM_NOME_{indice + 1}"
    texto = unicodedata.normalize("NFKD", str(nome).strip())
    texto = texto.encode("ascii", "ignore").decode("ascii").upper()
    texto = re.sub(r"[^A-Z0-9]+", "_", texto).strip("_")
    if texto and texto[0].isdigit():
        texto = f"COL_{texto}"
    return texto or f"COLUNA_SEM_NOME_{indice + 1}"


def nomes_unicos(cabecalho_original):
    """Evita perda silenciosa caso a fonte publique cabeçalhos repetidos."""
    usados = {}
    resultado = []
    for indice, nome in enumerate(cabecalho_original):
        base = normalizar_identificador(nome, indice)
        usados[base] = usados.get(base, 0) + 1
        sufixo = "" if usados[base] == 1 else f"_{usados[base]}"
        resultado.append(f"{base}{sufixo}")
    return resultado


def normalizar_cod_ibge(valor):
    """Normaliza representações equivalentes como 3552205 e 3552205.0."""
    texto = renderizar_valor(valor)
    if texto is None:
        return None
    texto = texto.strip().replace(",", ".")
    try:
        decimal = Decimal(texto)
        if decimal == decimal.to_integral_value():
            return str(int(decimal))
    except InvalidOperation:
        pass
    return texto


def sha256_arquivo(caminho):
    digest = hashlib.sha256()
    with open(caminho, "rb") as arquivo:
        for bloco in iter(lambda: arquivo.read(1024 * 1024), b""):
            digest.update(bloco)
    return digest.hexdigest()


def hash_linha(cabecalho_original, valores_originais):
    """Hash somente do conteúdo fonte, sem arquivo, guia ou data de ingestão."""
    documento = {
        "cabecalho": [renderizar_valor(valor) for valor in cabecalho_original],
        "valores": valores_originais,
    }
    serializado = json.dumps(
        documento,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return hashlib.sha256(serializado.encode("utf-8")).hexdigest()


DT_INGESTAO = dt.datetime.now(dt.timezone.utc).replace(tzinfo=None)
TABELA_STAGING = "_bronze_recorte_sorocaba_staging"
spark.sql(f"DROP TABLE IF EXISTS {TABELA_STAGING}")


def processar_guia(caminho_xlsx, guia, ano):
    """Lê uma guia estadual e grava, em lotes, somente as linhas de Sorocaba."""
    workbook = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    try:
        if guia not in workbook.sheetnames:
            raise ValueError(
                f"Guia esperada '{guia}' ausente em {os.path.basename(caminho_xlsx)}. "
                f"Guias encontradas: {workbook.sheetnames}"
            )

        worksheet = workbook[guia]
        linhas = worksheet.iter_rows(values_only=True)
        try:
            cabecalho_original = list(next(linhas))
        except StopIteration as exc:
            raise ValueError(f"Guia vazia: {guia}") from exc

        colunas_fonte = nomes_unicos(cabecalho_original)
        candidatos_ibge = [
            nome for nome in ("CD_IBGE", "COD_IBGE") if nome in colunas_fonte
        ]
        if not candidatos_ibge:
            raise ValueError(
                f"A coluna CD_IBGE/COD_IBGE não foi encontrada em {guia}. "
                f"Cabeçalhos normalizados: {colunas_fonte}"
            )
        indice_ibge = colunas_fonte.index(candidatos_ibge[0])

        schema_lote = StructType(
            [StructField(nome, StringType(), True) for nome in colunas_fonte]
            + [
                StructField("id_registro_fonte", StringType(), False),
                StructField("_arquivo_origem", StringType(), False),
                StructField("_guia_origem", StringType(), False),
                StructField("_ano_arquivo", IntegerType(), False),
                StructField("_dt_ingestao", TimestampType(), False),
            ]
        )

        nome_arquivo = os.path.basename(caminho_xlsx)
        buffer_sorocaba = []
        linhas_estaduais = 0
        linhas_sorocaba = 0

        def gravar_lote():
            nonlocal linhas_sorocaba
            if not buffer_sorocaba:
                return
            df_lote = spark.createDataFrame(buffer_sorocaba, schema=schema_lote)
            (
                df_lote.write.format("delta")
                .mode("append")
                .option("mergeSchema", "true")
                .saveAsTable(TABELA_STAGING)
            )
            linhas_sorocaba += len(buffer_sorocaba)
            buffer_sorocaba.clear()
            print(f"      {linhas_sorocaba:,} linhas de Sorocaba gravadas")

        for linha in linhas:
            valores = list(linha[: len(colunas_fonte)])
            if len(valores) < len(colunas_fonte):
                valores.extend([None] * (len(colunas_fonte) - len(valores)))
            if not any(valor is not None for valor in valores):
                continue

            linhas_estaduais += 1
            if normalizar_cod_ibge(valores[indice_ibge]) != COD_IBGE_SOROCABA:
                continue

            valores_texto = [renderizar_valor(valor) for valor in valores]
            id_registro = hash_linha(cabecalho_original, valores_texto)
            buffer_sorocaba.append(
                tuple(valores_texto)
                + (id_registro, nome_arquivo, guia, ano, DT_INGESTAO)
            )
            if len(buffer_sorocaba) >= LOTE_SOROCABA:
                gravar_lote()

        gravar_lote()
        return linhas_estaduais, linhas_sorocaba, colunas_fonte
    finally:
        workbook.close()

# COMMAND ----------
# MAGIC %md
# MAGIC ## 3. Reconstrução do recorte Bronze e do manifesto

# COMMAND ----------

manifesto = []
colunas_fonte_encontradas = set()

for ano, (arquivo, guias) in CONFIG_ANOS.items():
    caminho = os.path.join(DIRETORIO_XLSX, arquivo)
    print(f"\n[{ano}] {arquivo}")
    total_estadual = 0
    total_sorocaba = 0

    for guia in guias:
        print(f"  Processando guia {guia} ...")
        n_estadual, n_sorocaba, colunas_guia = processar_guia(caminho, guia, ano)
        total_estadual += n_estadual
        total_sorocaba += n_sorocaba
        colunas_fonte_encontradas.update(colunas_guia)
        print(
            f"    {n_estadual:,} linhas estaduais lidas; "
            f"{n_sorocaba:,} linhas de Sorocaba mantidas"
        )

    manifesto.append(
        (
            ano,
            arquivo,
            URL_TEMPLATE.format(ano=ano),
            caminho,
            os.path.getsize(caminho),
            sha256_arquivo(caminho),
            ", ".join(guias),
            total_estadual,
            total_sorocaba,
            DT_INGESTAO,
            "OK",
        )
    )

if not spark.catalog.tableExists(TABELA_STAGING):
    raise RuntimeError(
        "Nenhuma linha de Sorocaba foi localizada. Confira o código IBGE e os arquivos fonte."
    )

spark.sql("""
    CREATE OR REPLACE TABLE bronze_recorte_sorocaba
    USING DELTA
    AS SELECT * FROM _bronze_recorte_sorocaba_staging
""")
spark.sql(f"DROP TABLE IF EXISTS {TABELA_STAGING}")

schema_manifesto = StructType(
    [
        StructField("ano_arquivo", IntegerType(), False),
        StructField("arquivo", StringType(), False),
        StructField("url_fonte", StringType(), False),
        StructField("caminho_volume", StringType(), False),
        StructField("tamanho_bytes", StringType(), False),
        StructField("sha256", StringType(), False),
        StructField("guias", StringType(), False),
        StructField("linhas_estaduais_lidas", StringType(), False),
        StructField("linhas_sorocaba_mantidas", StringType(), False),
        StructField("dt_ingestao", TimestampType(), False),
        StructField("status", StringType(), False),
    ]
)

# Campos numéricos nascem como texto e são convertidos explicitamente para BIGINT,
# evitando inferências diferentes entre versões do runtime.
df_manifesto = spark.createDataFrame(
    [
        (
            ano,
            arquivo,
            url,
            caminho,
            str(tamanho),
            checksum,
            guias,
            str(n_estadual),
            str(n_sorocaba),
            instante,
            status,
        )
        for (
            ano,
            arquivo,
            url,
            caminho,
            tamanho,
            checksum,
            guias,
            n_estadual,
            n_sorocaba,
            instante,
            status,
        ) in manifesto
    ],
    schema=schema_manifesto,
).select(
    "ano_arquivo",
    "arquivo",
    "url_fonte",
    "caminho_volume",
    F.col("tamanho_bytes").cast("long").alias("tamanho_bytes"),
    "sha256",
    "guias",
    F.col("linhas_estaduais_lidas").cast("long").alias("linhas_estaduais_lidas"),
    F.col("linhas_sorocaba_mantidas").cast("long").alias("linhas_sorocaba_mantidas"),
    "dt_ingestao",
    "status",
)

(
    df_manifesto.write.format("delta")
    .mode("overwrite")
    .option("overwriteSchema", "true")
    .saveAsTable("bronze_manifesto")
)

# COMMAND ----------
# MAGIC %md
# MAGIC ## 4. Metadados do Unity Catalog e evidências da carga

# COMMAND ----------

def literal_sql(texto):
    return "'" + texto.replace("'", "''") + "'"


spark.sql(
    "COMMENT ON TABLE bronze_recorte_sorocaba IS "
    + literal_sql(
        "Captura fiel dos campos originais da SSP-SP para o código IBGE 3552205; "
        "valores fonte preservados como texto e acrescidos apenas de auditoria."
    )
)
spark.sql(
    "COMMENT ON TABLE bronze_manifesto IS "
    + literal_sql(
        "Manifesto dos quatro XLSX originais carregados manualmente no Unity Catalog Volume."
    )
)

comentarios_manifesto = {
    "ano_arquivo": "Ano nominal do arquivo da SSP-SP.",
    "arquivo": "Nome do XLSX original no Volume.",
    "url_fonte": "URL pública de origem do arquivo.",
    "caminho_volume": "Caminho do arquivo original preservado no Volume.",
    "tamanho_bytes": "Tamanho do arquivo original em bytes.",
    "sha256": "Checksum SHA-256 calculado sobre o arquivo original.",
    "guias": "Guias de dados lidas no arquivo.",
    "linhas_estaduais_lidas": "Linhas de dados estaduais percorridas, sem o cabeçalho.",
    "linhas_sorocaba_mantidas": "Linhas cujo código IBGE da fonte é 3552205.",
    "dt_ingestao": "Instante UTC da reconstrução integral da Bronze.",
    "status": "Resultado da ingestão do arquivo.",
}
for coluna, comentario in comentarios_manifesto.items():
    spark.sql(
        f"ALTER TABLE bronze_manifesto ALTER COLUMN `{coluna}` COMMENT "
        + literal_sql(comentario)
    )

for coluna in spark.table("bronze_recorte_sorocaba").columns:
    if coluna == "id_registro_fonte":
        comentario = "SHA-256 determinístico do cabeçalho e valores da linha original, sem auditoria."
    elif coluna == "_arquivo_origem":
        comentario = "XLSX original que forneceu a linha."
    elif coluna == "_guia_origem":
        comentario = "Guia original que forneceu a linha."
    elif coluna == "_ano_arquivo":
        comentario = "Ano nominal do XLSX original."
    elif coluna == "_dt_ingestao":
        comentario = "Instante UTC da ingestão no recorte Bronze."
    else:
        comentario = (
            f"Valor textual preservado da coluna fonte {coluna}; o nome foi normalizado "
            "apenas para compatibilidade com Delta Lake."
        )
    spark.sql(
        f"ALTER TABLE bronze_recorte_sorocaba ALTER COLUMN `{coluna}` COMMENT "
        + literal_sql(comentario)
    )

total_bronze = spark.table("bronze_recorte_sorocaba").count()
total_manifesto = spark.sql(
    "SELECT SUM(linhas_sorocaba_mantidas) AS total FROM bronze_manifesto"
).first()["total"]
if total_bronze != total_manifesto:
    raise AssertionError(
        f"Contagem Bronze ({total_bronze}) difere do manifesto ({total_manifesto})."
    )

print(f"\nBronze concluída: {total_bronze:,} linhas de Sorocaba")
print(f"Colunas fonte preservadas: {len(colunas_fonte_encontradas)}")
display(spark.table("bronze_manifesto").orderBy("ano_arquivo"))
display(
    spark.sql("""
        SELECT _ano_arquivo AS ano, _arquivo_origem AS arquivo,
               _guia_origem AS guia, COUNT(*) AS linhas
        FROM bronze_recorte_sorocaba
        GROUP BY _ano_arquivo, _arquivo_origem, _guia_origem
        ORDER BY ano, guia
    """)
)
